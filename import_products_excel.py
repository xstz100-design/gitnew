"""
批量从供应商 Excel 导入商品到系统
Excel 格式:
  Row 1: Supplier Full name : XXXX
  Row 2: 主标题（合并单元格）
  Row 3: 子标题
  Row 4+: 数据行

用法:
  python import_products_excel.py <excel文件路径> [--url https://khmerai.cn] [--user admin] [--pass 密码]
"""

import sys
import re
import argparse
import openpyxl
import requests
import json

# ────────── 列索引（0-based，Excel col 1 = index 0）──────────
C_NO           = 0   # 编号
C_BARCODE_CASE = 1   # 外箱条码
C_BARCODE_PACK = 2   # 中包条码
C_BARCODE_UNIT = 3   # 单品条码
C_BRAND        = 4   # 品牌
C_NAME_EN      = 5   # 英文名
C_NAME_KH      = 6   # 高棉文名
C_NAME_ZH      = 7   # 中文名
C_PICTURE      = 8   # 图片（跳过）
C_UNIT_WEIGHT  = 9   # 单品重量，如 "20g"
C_PACK_FORMAT  = 10  # 包装规格，如 "1*20*20包"
C_INNER_PACK   = 11  # Inner pack/Case
C_UNIT_INNER   = 12  # Unit/Inner Pack
C_UNIT_CASE    = 13  # Unit/Case
C_COST_CASE    = 14  # Cost/Case
C_DC_PCT       = 15  # %DC
C_NET_CASE     = 16  # Net Cost/Case
C_NET_UNIT     = 17  # Net Cost/Unit
C_PRICE_INCL   = 18  # Price Incl VAT
C_PRICE_EXCL   = 19  # Price Excl VAT
C_GP_PCT       = 20  # % GP
C_SHELF_LIFE   = 21  # Shelf life (Days)
C_PRINCIPLE    = 22  # Principle company
C_COUNTRY      = 23  # Country of origin
# 最小包 (unit)
C_UNIT_W       = 24  # 宽度
C_UNIT_L       = 25  # 长度
C_UNIT_H       = 26  # 高度
C_UNIT_WKG     = 27  # 重量 kg
# 中包 (pack)
C_PACK_W       = 28
C_PACK_L       = 29
C_PACK_H       = 30
C_PACK_WKG     = 31
# 外箱 (case)
C_CASE_W       = 32
C_CASE_L       = 33
C_CASE_H       = 34
C_CASE_WKG     = 35


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


def to_int(v):
    if v is None:
        return None
    try:
        return int(float(str(v)))
    except Exception:
        return None


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_unit_weight(raw):
    """'20g' -> (20.0, 'G'), '500ml' -> (500.0, 'ML'), '1pcs' -> (1.0, 'Pcs')"""
    if not raw:
        return None, None
    raw = str(raw).strip()
    m = re.match(r'^([\d.]+)\s*([a-zA-Z]+)', raw)
    if not m:
        return None, None
    val = float(m.group(1))
    unit = m.group(2).upper()
    if unit in ('G', 'GRAM', 'GRAMS'):
        unit = 'G'
    elif unit in ('ML', 'CC'):
        unit = 'ML'
    else:
        unit = 'Pcs'
    return val, unit


def parse_shelf_life(raw):
    """'180Days' -> 180"""
    if raw is None:
        return None
    m = re.match(r'(\d+)', str(raw))
    return int(m.group(1)) if m else None


def parse_gp_percent(v):
    """0.28 -> 28.0  (如果已经是百分比形式如 28 则直接返回)"""
    f = to_float(v)
    if f is None:
        return None
    # 如果值 < 1，认为是小数形式 0.28 -> 28%
    if f < 1:
        return round(f * 100, 2)
    return f


def row_to_product(row_vals, supplier_name):
    """将一行数据转换成 API 请求体 dict"""
    def g(idx):
        if idx < len(row_vals):
            return row_vals[idx]
        return None

    # 必填字段
    name_zh = to_str(g(C_NAME_ZH))
    name_en = to_str(g(C_NAME_EN))
    name = name_zh or name_en
    if not name:
        return None, "name 为空，跳过"

    # price_usd 使用 net_cost_per_unit
    price_usd = to_float(g(C_NET_UNIT))
    if not price_usd or price_usd <= 0:
        price_usd = to_float(g(C_PRICE_EXCL))
    if not price_usd or price_usd <= 0:
        return None, "price_usd 无效，跳过"

    uw_val, uw_unit = parse_unit_weight(g(C_UNIT_WEIGHT))

    product = {
        "name": name,
        "name_en": name_en,
        "name_kh": to_str(g(C_NAME_KH)),
        "brand": to_str(g(C_BRAND)),
        "barcode": to_str(g(C_BARCODE_UNIT)) or to_str(g(C_BARCODE_CASE)),
        "unit": "件",
        "price_usd": price_usd,
        "retail_price_usd": to_float(g(C_PRICE_INCL)),
        "stock": 0,
        "stock_warning": 0,
        "is_active": True,
        "is_featured": False,
        # 供应商
        "supplier_name": supplier_name,
        "principle_company": to_str(g(C_PRINCIPLE)),
        "country_of_origin": to_str(g(C_COUNTRY)),
        # 基础
        "unit_weight_value": uw_val,
        "unit_weight_unit": uw_unit,
        "packing_format": to_str(g(C_PACK_FORMAT)),
        "gp_percent": parse_gp_percent(g(C_GP_PCT)),
        "shelf_life_days": parse_shelf_life(g(C_SHELF_LIFE)),
        # 包装层级
        "inner_pack_per_case": to_int(g(C_INNER_PACK)),
        "unit_per_inner_pack": to_int(g(C_UNIT_INNER)),
        "unit_per_case": to_int(g(C_UNIT_CASE)),
        # 成本
        "cost_per_case": to_float(g(C_COST_CASE)),
        "dc_percent": to_float(g(C_DC_PCT)),
        "net_cost_per_case": to_float(g(C_NET_CASE)),
        "net_cost_per_unit": to_float(g(C_NET_UNIT)),
        "price_incl_vat": to_float(g(C_PRICE_INCL)),
        "price_excl_vat": to_float(g(C_PRICE_EXCL)),
        # 尺寸 最小包
        "unit_width_cm": to_float(g(C_UNIT_W)),
        "unit_length_cm": to_float(g(C_UNIT_L)),
        "unit_height_cm": to_float(g(C_UNIT_H)),
        "unit_weight_kg": to_float(g(C_UNIT_WKG)),
        # 尺寸 中包
        "pack_width_cm": to_float(g(C_PACK_W)),
        "pack_length_cm": to_float(g(C_PACK_L)),
        "pack_height_cm": to_float(g(C_PACK_H)),
        "pack_weight_kg": to_float(g(C_PACK_WKG)),
        # 尺寸 外箱
        "case_width_cm": to_float(g(C_CASE_W)),
        "case_length_cm": to_float(g(C_CASE_L)),
        "case_height_cm": to_float(g(C_CASE_H)),
        "case_weight_kg": to_float(g(C_CASE_WKG)),
        # 排序
        "sort_order": to_int(g(C_NO)) or 0,
    }

    # 移除 None 值（API 可选字段不传）
    return {k: v for k, v in product.items() if v is not None}, None


def login(base_url, username, password):
    resp = requests.post(f"{base_url}/api/auth/login",
                         json={"username": username, "password": password},
                         timeout=15)
    resp.raise_for_status()
    data = resp.json()
    token = data.get("token") or data.get("access_token")
    if not token:
        raise ValueError(f"登录响应中无 token: {data}")
    return token


def create_product(base_url, token, product_data):
    resp = requests.post(
        f"{base_url}/api/products",
        json=product_data,
        headers={"Authorization": f"Bearer {token}"},
        timeout=15
    )
    try:
        return resp.status_code, resp.json()
    except Exception:
        return resp.status_code, {"detail": resp.text[:300]}


def main():
    parser = argparse.ArgumentParser(description="从供应商 Excel 批量导入商品")
    parser.add_argument("excel", help="Excel 文件路径")
    parser.add_argument("--url", default="https://khmerai.cn", help="API 基础地址")
    parser.add_argument("--user", default="admin", help="管理员用户名")
    parser.add_argument("--pass", dest="password", default="admin123", help="管理员密码")
    parser.add_argument("--dry-run", action="store_true", help="仅解析不实际上传")
    args = parser.parse_args()

    # 读取 Excel
    print(f"读取文件: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    ws = wb.active
    print(f"  工作表: {ws.title}，行数: {ws.max_row}，列数: {ws.max_column}")

    # 提取供应商名称（Row 1, Col 1）
    supplier_raw = ws.cell(1, 1).value or ""
    supplier_name = None
    m = re.search(r':\s*(.+)$', str(supplier_raw).strip())
    if m:
        supplier_name = m.group(1).strip()
    print(f"  供应商: {supplier_name}")

    # 收集数据行（从第4行开始，跳过空行）
    products_to_create = []
    skipped = []

    for row_idx in range(4, ws.max_row + 1):
        row_vals = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        # 跳过完全空行
        if all(v is None for v in row_vals):
            continue
        # 跳过 No. 为空且名称也为空的行
        if row_vals[C_NO] is None and row_vals[C_NAME_EN] is None and row_vals[C_NAME_ZH] is None:
            continue

        product, err = row_to_product(row_vals, supplier_name)
        if err:
            skipped.append(f"  第 {row_idx} 行: {err}")
            continue
        products_to_create.append((row_idx, product))

    print(f"\n解析完成: 有效商品 {len(products_to_create)} 条，跳过 {len(skipped)} 条")
    if skipped:
        for s in skipped:
            print(s)

    if not products_to_create:
        print("没有可导入的商品，退出")
        return

    # 预览
    print("\n─── 待导入商品预览 ───")
    for row_idx, p in products_to_create:
        print(f"  [{row_idx}] {p.get('name')} | 条码={p.get('barcode')} | 价格={p.get('price_usd')} | 供应商={p.get('supplier_name')}")

    if args.dry_run:
        print("\n[dry-run 模式] 已跳过实际上传")
        print("\n完整数据（第一条）:")
        print(json.dumps(products_to_create[0][1], ensure_ascii=False, indent=2))
        return

    # 登录
    print(f"\n登录 {args.url} ...")
    try:
        token = login(args.url, args.user, args.password)
        print("  登录成功")
    except Exception as e:
        print(f"  登录失败: {e}")
        sys.exit(1)

    # 批量上传
    print(f"\n开始上传 {len(products_to_create)} 条商品 ...")
    success, failed = 0, 0
    for row_idx, product in products_to_create:
        status, resp = create_product(args.url, token, product)
        if status in (200, 201):
            prod_id = resp.get("id", "?")
            print(f"  ✓ [{row_idx}] {product['name']}  (id={prod_id})")
            success += 1
        else:
            detail = resp.get("detail") or resp.get("message") or str(resp)
            print(f"  ✗ [{row_idx}] {product['name']}  HTTP {status}: {detail}")
            failed += 1

    print(f"\n完成: 成功 {success} 条，失败 {failed} 条")


if __name__ == "__main__":
    main()
